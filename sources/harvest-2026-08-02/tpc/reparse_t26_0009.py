"""Coordinate-aware re-parse of TPC T26-0009 (OBBBA by racial/ethnic group).

The overnight harvest's generic distribution parser (parse_tpc.py, kept
with the raw downloads outside the repo) mis-staged this workbook: each of
its five sheets (All Tax Units + four racial/ethnic groups) stacks TWO
panels — a federal-tax-CHANGE panel (rows 15-25) above a BASELINE
income/tax distribution panel (rows 37-47) — that reuse the same income
group labels and column letters. The generic parser swept every
group-labeled row under one header, so baseline-panel dollar cells landed
in percent metrics (e.g. All!K37 = 1020, the average baseline federal tax
burden in dollars, staged as a percent change in after-tax income), and
the race axis (sheet identity) was collapsed entirely.

This parser is coordinate-pinned per sheet and stages ONLY the tax-change
panel's mission columns, race as a subgroup condition:

    C  Tax Units with Tax Increase or Cut / With Tax Cut / Pct of Tax
       Units                               -> share_with_tax_cut (percent)
    K  Percent Change in After-Tax Income  -> pct_change_after_tax_income
                                              (percent)
    O  Average Federal Tax Change ($)      -> avg_tax_change_usd
                                              (USD per tax unit)

5 sheets x 9 income groups x 3 metrics = 135 rows. The baseline panel
(income/tax LEVELS under pre-OBBBA law) stays unstaged, same as the
harvest's other unstaged baseline-level tables (see NOTES.md "Unstaged
downloads"). Every anchor cell (title block, headers, row labels) is
asserted verbatim before any value is read — drift fails loudly.

Usage (rewrites claims_staged.jsonl in place, replacing the 48 defective
T26-0009 rows; verifies the workbook sha256 against manifest.jsonl and
re-verifies the spliced file cell-by-cell afterwards):

    uv run --with openpyxl python reparse_t26_0009.py [workbook.xlsx]

The workbook default is the harvest download tree
(~/scorecard-harvest/tpc/downloads/T26-0009.xlsx); pass a path when the
raw tree lives elsewhere.
"""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
TABLE_ID = "T26-0009"

# sheet name -> (population line at A7/A30, subgroup condition slug;
# None = the all-tax-units sheet, which stages no subgroup key)
SHEETS = {
    "All": ("All Tax Units", None),
    "White": ("White, Non-Hispanic", "white_non_hispanic"),
    "Black": ("Black, Non-Hispanic", "black_non_hispanic"),
    "Hispanic": ("Hispanic", "hispanic"),
    "Additional Races": ("Additional Races", "additional_races"),
}

TITLE_BLOCK = {
    "A3": "Table T26-0009",
    "A4": "Tax Provisions in Title VII of H.R.1",
    "A5": "Baseline: Law Prior to the 2025 Budget Reconciliation Act",
    "A6": "Distribution of Federal Tax Change by Expanded Cash Income"
          " Percentile, 2026 ¹",
}

# Tax-change panel header anchors (identical on all five sheets).
HEADER_ANCHORS = {
    "A9": "Expanded Cash Income Percentile 2,3",
    "C9": "Tax Units with Tax Increase or Cut 4",
    "K9": "Percent Change in After-Tax Income 5",
    "M9": "Share of Total Federal Tax Change",
    "O9": "Average Federal Tax Change ($)",
    "Q9": "Average Federal Tax Rate 6",
    "C11": "With Tax Cut",
    "G11": "With Tax Increase",
    "C12": "Pct of Tax Units",
    "E12": "Avg Tax Change ($)",
    # Baseline-panel anchors: assert the second panel is where we think it
    # is (and therefore OUTSIDE the rows read below).
    "A28": "Baseline Distribution of Income and Federal Taxes",
    "A32": "Expanded Cash Income Percentile 2,3",
    "C32": "Tax Units",
}

# Tax-change panel data rows (baseline panel = rows 37-47, never read).
DATA_ROWS = {
    15: "Lowest Quintile",
    16: "Second Quintile",
    17: "Middle Quintile",
    18: "Fourth Quintile",
    19: "Top Quintile",
    20: "All",
    23: "80-90",
    24: "90-95",
    25: "Top 5 Percent",
}

# column -> (metric, unit, dollar-valued?, source_column composite in the
# harvest's lowercased "main / sub / sub" header style)
COLUMNS = {
    "C": (
        "share_with_tax_cut", "percent_of_tax_units", False,
        "tax units with tax increase or cut 4 / with tax cut /"
        " pct of tax units",
    ),
    "K": (
        "pct_change_after_tax_income", "percent", False,
        "percent change in after-tax income 5",
    ),
    "O": (
        "avg_tax_change_usd", "usd_per_tax_unit", True,
        "average federal tax change ($)",
    ),
}


def manifest_entry() -> dict:
    for line in (HERE / "manifest.jsonl").open():
        row = json.loads(line)
        if row.get("table_id") == TABLE_ID:
            return row
    raise ValueError(f"{TABLE_ID} missing from manifest.jsonl")


def check_anchor(ws, coord: str, expected: str) -> None:
    got = ws[coord].value
    if got != expected:
        raise ValueError(
            f"{ws.title}!{coord}: expected {expected!r}, found {got!r}"
        )


def parse(workbook: Path) -> list[dict]:
    manifest = manifest_entry()
    sha = hashlib.sha256(workbook.read_bytes()).hexdigest()
    if sha != manifest["sha256"]:
        raise ValueError(
            f"workbook sha256 {sha} != manifested {manifest['sha256']}"
        )

    wb = openpyxl.load_workbook(workbook, data_only=True)
    if wb.sheetnames != list(SHEETS):
        raise ValueError(f"sheet drift: {wb.sheetnames}")

    rows = []
    for sheet, (population, subgroup) in SHEETS.items():
        ws = wb[sheet]
        for coord, expected in TITLE_BLOCK.items():
            check_anchor(ws, coord, expected)
        for coord, expected in HEADER_ANCHORS.items():
            check_anchor(ws, coord, expected)
        check_anchor(ws, "A7", population)
        check_anchor(ws, "A30", population)
        for r, label in DATA_ROWS.items():
            check_anchor(ws, f"A{r}", label)

        for col, (metric, unit, dollars, source_column) in COLUMNS.items():
            for r, group in DATA_ROWS.items():
                cell = f"{col}{r}"
                raw = ws[cell].value
                if not isinstance(raw, (int, float)):
                    raise ValueError(
                        f"{sheet}!{cell}: non-numeric {raw!r} — suppressed"
                        " cells are unexpected in this workbook"
                    )
                # These sheets store the display numbers themselves
                # (integer dollars, one-decimal percents) — no unrounded
                # twin block exists, so hidden precision means the layout
                # changed under us.
                if dollars and raw != int(raw):
                    raise ValueError(f"{sheet}!{cell}: non-integer dollars")
                if not dollars and round(raw, 1) != raw:
                    raise ValueError(f"{sheet}!{cell}: >1dp percent")

                conditions = {
                    "geography": "US",
                    "income_group": group,
                    "income_axis": HEADER_ANCHORS["A9"],
                }
                if subgroup is not None:
                    conditions["subgroup"] = subgroup
                rows.append(
                    {
                        "proposed_metric": metric,
                        "proposed_unit": unit,
                        "value": float(raw),
                        "conditions": conditions,
                        "source_column": source_column,
                        "period": 2026,
                        "time_basis": "annual",
                        "source": "tpc",
                        "source_model": "tpc_microsim",
                        "table_id": TABLE_ID,
                        "reform_hint": TITLE_BLOCK["A4"],
                        "baseline_hint": TITLE_BLOCK["A5"],
                        "calibration_relationship": "held_out",
                        "publication": {
                            "title": manifest["title"],
                            "url": manifest["url"],
                            "date": manifest["date"],
                            "vintage":
                                "TPC model estimates T-series (PRELIMINARY)",
                            "file": manifest["local_path"],
                            "sha256": manifest["sha256"],
                            "sheet": sheet,
                            "cell": cell,
                            "population": population,
                        },
                    }
                )
    if len(rows) != len(SHEETS) * len(DATA_ROWS) * len(COLUMNS):
        raise ValueError(f"expected 135 rows, built {len(rows)}")
    return rows


def splice(staged_path: Path, rows: list[dict]) -> tuple[int, int]:
    lines = staged_path.read_text().splitlines()
    keep, defective_at = [], None
    for i, line in enumerate(lines):
        if json.loads(line)["table_id"] == TABLE_ID:
            defective_at = i if defective_at is None else defective_at
        else:
            keep.append(line)
    n_defective = len(lines) - len(keep)
    if defective_at is None:
        defective_at = len(keep)
    new_lines = [json.dumps(r) for r in rows]
    out = keep[:defective_at] + new_lines + keep[defective_at:]
    staged_path.write_text("\n".join(out) + "\n")
    return n_defective, len(new_lines)


def verify(staged_path: Path, workbook: Path) -> int:
    """Independent post-write check: every T26-0009 row in the staged file
    must equal the workbook cell its publication names, coverage must be
    the full 5x9x3 cross-product, and no defective-era shape may survive
    (rows without a cell reference, or dollar magnitudes under percent
    units)."""
    wb = openpyxl.load_workbook(workbook, data_only=True)
    staged = [
        json.loads(line)
        for line in staged_path.open()
        if json.loads(line)["table_id"] == TABLE_ID
    ]
    if len(staged) != 135:
        raise ValueError(f"expected 135 staged rows, found {len(staged)}")
    seen = set()
    for row in staged:
        pub = row["publication"]
        cell_value = wb[pub["sheet"]][pub["cell"]].value
        if float(cell_value) != row["value"]:
            raise ValueError(
                f"{pub['sheet']}!{pub['cell']} = {cell_value},"
                f" staged {row['value']}"
            )
        if row["proposed_unit"] in ("percent", "percent_of_tax_units"):
            if not -10 <= row["value"] <= 100:
                raise ValueError(
                    f"percent row out of range: {row['value']}"
                    f" ({pub['sheet']}!{pub['cell']})"
                )
        seen.add(
            (
                pub["sheet"],
                row["conditions"]["income_group"],
                row["proposed_metric"],
                row["conditions"].get("subgroup"),
            )
        )
    expected = {
        (sheet, group, metric, subgroup)
        for sheet, (_, subgroup) in SHEETS.items()
        for group in DATA_ROWS.values()
        for metric, _, _, _ in COLUMNS.values()
    }
    if seen != expected:
        raise ValueError(
            f"coverage mismatch: missing {expected - seen},"
            f" extra {seen - expected}"
        )
    return len(staged)


def main() -> None:
    workbook = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else Path.home() / "scorecard-harvest/tpc/downloads/T26-0009.xlsx"
    )
    staged_path = HERE / "claims_staged.jsonl"
    rows = parse(workbook)
    removed, added = splice(staged_path, rows)
    verified = verify(staged_path, workbook)
    total = sum(1 for _ in staged_path.open())
    print(
        json.dumps(
            {
                "removed_defective": removed,
                "staged": added,
                "verified_against_workbook": verified,
                "claims_staged_total": total,
            }
        )
    )


if __name__ == "__main__":
    main()
